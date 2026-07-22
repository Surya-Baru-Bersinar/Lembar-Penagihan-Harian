import sys
import traceback
import datetime
import openpyxl
import gspread
from google.oauth2.service_account import Credentials

def normalize_text(text):
    if not text:
        return ""
    return " ".join(str(text).strip().lower().split())

def format_value(val):
    if val is None:
        return ""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime('%d/%m/%Y')
    return val

def baca_konfigurasi(filepath='piutang.conf'):
    config = {}
    current_section = None
    
    try:
        with open(filepath, 'r') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                
                if line.startswith('[') and line.endswith(']'):
                    current_section = line[1:-1].strip().upper()
                    if current_section not in config:
                        config[current_section] = {}
                else:
                    if current_section is not None:
                        if '=' in line:
                            k, v = line.split('=', 1)
                            config[current_section][k.strip().upper()] = v.strip()
                        else:
                            if 'VALUE' in config[current_section]:
                                config[current_section]['VALUE'] += '\n' + line
                            else:
                                config[current_section]['VALUE'] = line
    except Exception as e:
        print(f"--> Gagal membaca {filepath}: {e}")
        sys.exit(1)
        
    return config

config = baca_konfigurasi('piutang.conf')
perusahaan = config.get('PERUSAHAAN', {}).get('VALUE', '')
divisi = config.get('DIVISI', {}).get('VALUE', '')
tanggal = config.get('TANGGAL', {}).get('VALUE', '')
input_user = config.get('INPUT', {}).get('VALUE', '')

ss_url = config.get('SS', {}).get('URL', '')
target_sheet_name = config.get('SS', {}).get('SHEET_NAME', '')

try:
    wb = openpyxl.load_workbook('Print_AR_temp.xlsx', data_only=True)
    ws = wb.active
    
    data_to_insert = []
    for row in range(1, ws.max_row + 1):
        col_a = ws.cell(row=row, column=1).value
        if col_a is not None and str(col_a).strip() != "":
            baris_baru = [
                perusahaan,
                format_value(ws.cell(row=row, column=1).value),
                divisi,
                tanggal,
                input_user,
                format_value(ws.cell(row=row, column=2).value),
                format_value(ws.cell(row=row, column=3).value),
                format_value(ws.cell(row=row, column=4).value),
                format_value(ws.cell(row=row, column=5).value),
                format_value(ws.cell(row=row, column=6).value),
                format_value(ws.cell(row=row, column=7).value),
                format_value(ws.cell(row=row, column=8).value),
                format_value(ws.cell(row=row, column=9).value),
                format_value(ws.cell(row=row, column=10).value)
            ]
            data_to_insert.append(baris_baru)
            
    if not data_to_insert:
        print("--> Tidak ada data yang ditemukan untuk disisipkan.")
        sys.exit()
        
    scopes = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive'
    ]
    
    credentials = Credentials.from_service_account_file('credentials.json', scopes=scopes)
    gc = gspread.authorize(credentials)
    
    if not ss_url:
        print("--> Error: URL Google Spreadsheet pada section [SS] tidak ditemukan di piutang.conf")
        sys.exit()

    try:
        sh = gc.open_by_url(ss_url)
    except gspread.exceptions.SpreadsheetNotFound:
        print("--> Error: Spreadsheets tidak ditemukan. Pastikan URL benar dan email Service Account sudah diberi akses Editor.")
        sys.exit()
    except Exception as e:
        print(f"--> Error saat membuka URL Spreadsheet: {e}")
        sys.exit()
        
    target_norm = normalize_text(target_sheet_name)
    
    worksheet = None
    
    for ws_item in sh.worksheets():
        if normalize_text(ws_item.title) == target_norm:
            worksheet = ws_item
            break
            
    if worksheet is None:
        print(f"--> Error: Sheet '{target_sheet_name}' tidak ditemukan.")
        print("    Daftar Sheet yang tersedia di Google Sheets:", [w.title for w in sh.worksheets()])
        sys.exit()
    
    semua_nilai = worksheet.get_all_values()
    total_baris = len(semua_nilai)
    
    baris_sisip = total_baris - 1
    if baris_sisip < 1:
        baris_sisip = 1
        
    print(f"--> Menyiapkan {len(data_to_insert)} baris untuk disisipkan pada baris ke-{baris_sisip} di Sheet '{worksheet.title}'...")
    
    worksheet.insert_rows(data_to_insert, row=baris_sisip, value_input_option='USER_ENTERED', inherit_from_before=True)
    
    print("--> Proses selesai, data berhasil disisipkan ke Spreadsheets dengan format bawaan.")
    
except Exception as e:
    print(f"--> Error detail:\n{traceback.format_exc()}")