import openpyxl

filename = "ExportFile_clean_temp.xlsx"
wb = openpyxl.load_workbook(filename)
ws = wb.active

col_idx = None
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    if val and str(val).strip() == "Sisa Piutang":
        col_idx = col
        break

deleted_count = 0

if col_idx:
    for row in range(ws.max_row, 1, -1):
        val = ws.cell(row=row, column=col_idx).value
        if val is not None:
            val_str = str(val).strip().replace(".", "").replace(",", ".")
            try:
                if float(val_str) == 0.0:
                    ws.delete_rows(row)
                    deleted_count += 1
            except ValueError:
                pass

    wb.save(filename)
    print(f"--> Proses selesai. Berhasil menghapus {deleted_count} baris dengan Sisa Piutang 0.")
else:
    print("--> Kolom Sisa Piutang tidak ditemukan.")