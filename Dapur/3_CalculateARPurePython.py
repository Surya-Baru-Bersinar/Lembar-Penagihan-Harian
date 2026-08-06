import copy
from datetime import datetime
import sys
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd

config = {}
pure_config = {}
current_key = None

try:
    with open("piutang.conf", "r") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            if line.startswith("[") and line.endswith("]"):
                current_key = line[1:-1]
            else:
                if current_key and current_key not in config:
                    config[current_key] = line

                if current_key == "PURE" and "=" in line:
                    k, v = line.split("=", 1)
                    pure_config[k.strip()] = v.strip()
except Exception as e:
    print(f"--> Gagal membaca piutang.conf: {e}")

if pure_config.get("pr_process", "").strip().lower() != "ya":
    print(
        "--> Opsi pr_process pada [PURE] bernilai No. Proses dilewati"
    )
    sys.exit()

tgl_config_str = config.get("TANGGAL", "")
tgl_config_val = tgl_config_str

if tgl_config_str:
    try:
        tgl_config_val = datetime.strptime(tgl_config_str, "%d/%m/%Y").date()
    except ValueError:
        tgl_config_val = tgl_config_str

df = pd.read_excel("Laporan_Piutang_Penagih_temp.xlsx")

if "Tgl Faktur" in df.columns:
    df["Tgl Faktur"] = pd.to_datetime(
        df["Tgl Faktur"], dayfirst=True, errors="coerce"
    )

df_data = df[df["No"].notna()].copy()


def copy_range(
    ws_src, start_row, end_row, ws_dst, target_start_row, max_col=16
):
    row_offset = target_start_row - start_row

    for r in range(start_row, end_row + 1):
        dst_r = r + row_offset

        if r in ws_src.row_dimensions:
            ws_dst.row_dimensions[dst_r].height = ws_src.row_dimensions[
                r
            ].height

        for c in range(1, max_col + 1):
            cell_src = ws_src.cell(row=r, column=c)
            cell_dst = ws_dst.cell(row=dst_r, column=c)

            cell_dst.value = cell_src.value

            if cell_src.has_style:
                cell_dst.font = copy.copy(cell_src.font)
                cell_dst.border = copy.copy(cell_src.border)
                cell_dst.fill = copy.copy(cell_src.fill)
                cell_dst.number_format = cell_src.number_format
                cell_dst.protection = copy.copy(cell_src.protection)
                cell_dst.alignment = copy.copy(cell_src.alignment)

    for merged_range in list(ws_src.merged_cells.ranges):
        if (
            merged_range.min_row >= start_row
            and merged_range.max_row <= end_row
        ):
            new_min_r = merged_range.min_row + row_offset
            new_max_r = merged_range.max_row + row_offset
            ws_dst.merge_cells(
                start_row=new_min_r,
                start_column=merged_range.min_col,
                end_row=new_max_r,
                end_column=merged_range.max_col,
            )


try:
    wb = openpyxl.load_workbook("TEMPLATE.xlsm", keep_vba=True)

    ws_temp = wb.active
    ws_temp.title = "TEMP_DESIGN"

    max_row_temp = ws_temp.max_row
    if max_row_temp < 7:
        max_row_temp = 7

    ws_out = wb.create_sheet(title="Print AR")

    current_out_row = 1

    if "Halaman" in df_data.columns:
        groups = df_data.groupby(["Penagih", "Halaman"], sort=False)
    else:
        groups = df_data.groupby("Penagih", sort=False)

    for group_keys, group_df in groups:
        penagih = group_keys[0] if isinstance(group_keys, tuple) else group_keys
        start_row_for_group = current_out_row

        copy_range(ws_temp, 1, 4, ws_out, current_out_row)

        ws_out.cell(row=start_row_for_group + 1, column=4, value=penagih)
        ws_out.cell(
            row=start_row_for_group + 1,
            column=8,
            value=config.get("PERUSAHAAN", ""),
        )
        ws_out.cell(
            row=start_row_for_group + 1,
            column=11,
            value=config.get("DIVISI", ""),
        )
        ws_out.cell(
            row=start_row_for_group + 1, column=15, value=tgl_config_val
        )
        ws_out.cell(
            row=start_row_for_group + 2,
            column=15,
            value=config.get("INPUT", ""),
        )

        current_out_row += 4
        data_start_row = current_out_row

        for _, row_data in group_df.iterrows():
            copy_range(ws_temp, 5, 5, ws_out, current_out_row)

            ws_out.cell(
                row=current_out_row,
                column=2,
                value=int(row_data["No"]) if pd.notna(row_data["No"]) else "",
            )
            ws_out.cell(
                row=current_out_row,
                column=3,
                value=(
                    row_data.get("Kode", "")
                    if pd.notna(row_data.get("Kode"))
                    else ""
                ),
            )
            ws_out.cell(
                row=current_out_row,
                column=4,
                value=(
                    row_data.get("Nama Pelanggan", "")
                    if pd.notna(row_data.get("Nama Pelanggan"))
                    else ""
                ),
            )
            ws_out.cell(
                row=current_out_row,
                column=5,
                value=(
                    row_data.get("Umur JT", "")
                    if pd.notna(row_data.get("Umur JT"))
                    else ""
                ),
            )
            ws_out.cell(
                row=current_out_row,
                column=6,
                value=(
                    row_data.get("No. Faktur", "")
                    if pd.notna(row_data.get("No. Faktur"))
                    else ""
                ),
            )

            tgl_faktur = row_data.get("Tgl Faktur", None)
            ws_out.cell(
                row=current_out_row,
                column=7,
                value=(
                    tgl_faktur.strftime("%d/%m/%Y")
                    if pd.notna(tgl_faktur)
                    else ""
                ),
            )

            val_faktur = row_data.get("Nilai Faktur", None)
            val_terbayar = row_data.get("Terbayar", None)
            val_sisa = row_data.get("Sisa Piutang", None)

            ws_out.cell(
                row=current_out_row,
                column=8,
                value=(
                    val_faktur
                    if pd.notna(val_faktur) and val_faktur != ""
                    else None
                ),
            )
            ws_out.cell(
                row=current_out_row,
                column=9,
                value=(
                    val_terbayar
                    if pd.notna(val_terbayar) and val_terbayar != ""
                    else None
                ),
            )
            ws_out.cell(
                row=current_out_row,
                column=10,
                value=(
                    val_sisa if pd.notna(val_sisa) and val_sisa != "" else None
                ),
            )

            current_out_row += 1

        data_end_row = current_out_row - 1

        copy_range(ws_temp, 6, 6, ws_out, current_out_row)
        ws_out.cell(row=current_out_row, column=2, value="TOTAL TAGIHAN")
        ws_out.cell(
            row=current_out_row,
            column=8,
            value=f"=SUM(H{data_start_row}:H{data_end_row})",
        )
        ws_out.cell(
            row=current_out_row,
            column=9,
            value=f"=SUM(I{data_start_row}:I{data_end_row})",
        )
        ws_out.cell(
            row=current_out_row,
            column=10,
            value=f"=SUM(J{data_start_row}:J{data_end_row})",
        )

        current_out_row += 1

        if max_row_temp >= 7:
            jumlah_baris_footer = max_row_temp - 7 + 1
            copy_range(ws_temp, 7, max_row_temp, ws_out, current_out_row)
            current_out_row += jumlah_baris_footer

        current_out_row += 2

    lebar_spesifik = {
        'B': 10,'C': 16,'D': 75,'E': 15,'F': 30, 'G': 32, 'H': 35, 'I': 35, 'J': 35,
        'K': 25, 'L': 35, 'M': 15, 'N': 37, 'O': 37, 'P': 30
    }
    for col_letter, width in lebar_spesifik.items():
        ws_out.column_dimensions[col_letter].width = width

    for r in range(1, current_out_row + 1):
        for c in range(1, 17):
            val = ws_out.cell(row=r, column=c).value
            if isinstance(val, str) and "TTD SALES & COLLECTOR" in val:
                ws_out.row_dimensions[r].height = 115
                break

    wb.remove(ws_temp)

    wb.save("Print_AR.xlsm")
    print("--> Proses ekspor berhasil! File disimpan sebagai Print_AR.xlsm")

except Exception as e:
    print(f"--> Terjadi kesalahan saat memproses file Excel: {e}")