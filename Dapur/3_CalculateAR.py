import pandas as pd
import openpyxl
from copy import copy
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

config = {}
current_key = None
try:
    with open('piutang.conf', 'r') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            if line.startswith('[') and line.endswith(']'):
                current_key = line[1:-1]
            else:
                if current_key and current_key not in config:
                    config[current_key] = line
except Exception as e:
    print(f"--> Gagal membaca piutang.conf: {e}")

df = pd.read_excel('Laporan_Piutang_Penagih_temp.xlsx')

df = df[df['Penagih'].notna()]
df = df[~df['Penagih'].astype(str).str.startswith('TOTAL', na=False)]
df = df[df['Kode'].notna()]
df = df[df['Kode'].astype(str).str.strip() != '']

wb_temp = openpyxl.load_workbook('TEMPLATE.xlsx')
ws_temp = wb_temp.active

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Print AR"

for col_letter, col_dim in ws_temp.column_dimensions.items():
    ws_out.column_dimensions[col_letter].width = col_dim.width
    ws_out.column_dimensions[col_letter].hidden = col_dim.hidden

def copy_cell(source_cell, target_cell):
    if source_cell.data_type == 'f' and source_cell.value:
        target_cell.value = Translator(source_cell.value, source_cell.coordinate).translate_formula(target_cell.coordinate)
    else:
        target_cell.value = source_cell.value
        
    target_cell.data_type = source_cell.data_type
    
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

current_out_row = 1
groups = df.groupby('Penagih', sort=False)

for penagih, group_df in groups:
    start_row_for_group = current_out_row
    
    for r in range(1, 5):
        if r in ws_temp.row_dimensions and ws_temp.row_dimensions[r].height is not None:
            ws_out.row_dimensions[current_out_row].height = ws_temp.row_dimensions[r].height
        for c in range(1, ws_temp.max_column + 1):
            copy_cell(ws_temp.cell(row=r, column=c), ws_out.cell(row=current_out_row, column=c))
        current_out_row += 1
        
    ws_out.cell(row=start_row_for_group + 1, column=4).value = penagih
    ws_out.cell(row=start_row_for_group + 1, column=8).value = config.get('PERUSAHAAN', '')
    ws_out.cell(row=start_row_for_group + 1, column=11).value = config.get('DIVISI', '')
    ws_out.cell(row=start_row_for_group + 1, column=15).value = config.get('TANGGAL', '')
    ws_out.cell(row=start_row_for_group + 2, column=15).value = config.get('INPUT', '')
    
    for merged_range in ws_temp.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_row <= 4 and max_row <= 4:
            try:
                ws_out.merge_cells(
                    start_row=min_row + start_row_for_group - 1, 
                    start_column=min_col, 
                    end_row=max_row + start_row_for_group - 1, 
                    end_column=max_col
                )
            except Exception:
                pass
                
    data_start_row = current_out_row
    nomor = 1
    
    for _, row_data in group_df.iterrows():
        if 5 in ws_temp.row_dimensions and ws_temp.row_dimensions[5].height is not None:
            ws_out.row_dimensions[current_out_row].height = ws_temp.row_dimensions[5].height
        for c in range(1, ws_temp.max_column + 1):
            copy_cell(ws_temp.cell(row=5, column=c), ws_out.cell(row=current_out_row, column=c))
            
        ws_out.cell(row=current_out_row, column=2).value = nomor
        ws_out.cell(row=current_out_row, column=3).value = row_data.get('Kode', '')
        ws_out.cell(row=current_out_row, column=4).value = row_data.get('Nama Pelanggan', '')
        ws_out.cell(row=current_out_row, column=5).value = row_data.get('Umur JT', '')
        ws_out.cell(row=current_out_row, column=6).value = row_data.get('No. Faktur', '')
        ws_out.cell(row=current_out_row, column=7).value = row_data.get('Tgl Faktur', '')
        
        for col_idx, col_name in zip([8, 9, 10], ['Nilai Faktur', 'Terbayar', 'Sisa Piutang']):
            val = row_data.get(col_name, '')
            if pd.notna(val) and val != "":
                ws_out.cell(row=current_out_row, column=col_idx).value = val
            else:
                ws_out.cell(row=current_out_row, column=col_idx).value = None

        current_out_row += 1
        nomor += 1
        
    data_end_row = current_out_row - 1
    
    if 6 in ws_temp.row_dimensions and ws_temp.row_dimensions[6].height is not None:
        ws_out.row_dimensions[current_out_row].height = ws_temp.row_dimensions[6].height
    for c in range(1, ws_temp.max_column + 1):
        copy_cell(ws_temp.cell(row=6, column=c), ws_out.cell(row=current_out_row, column=c))
        
    try:
        ws_out.merge_cells(start_row=current_out_row, start_column=2, end_row=current_out_row, end_column=7)
    except Exception:
        pass
        
    ws_out.cell(row=current_out_row, column=2).value = "TOTAL TAGIHAN"
    ws_out.cell(row=current_out_row, column=8).value = f"=SUM(H{data_start_row}:H{data_end_row})"
    ws_out.cell(row=current_out_row, column=9).value = f"=SUM(I{data_start_row}:I{data_end_row})"
    ws_out.cell(row=current_out_row, column=10).value = f"=SUM(J{data_start_row}:J{data_end_row})"
    
    for merged_range in ws_temp.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_row == 6 and min_col != 2:
            try:
                ws_out.merge_cells(
                    start_row=current_out_row, 
                    start_column=min_col, 
                    end_row=current_out_row + (max_row - min_row), 
                    end_column=max_col
                )
            except Exception:
                pass

    current_out_row += 1

    if ws_temp.max_row >= 7:
        for r in range(7, ws_temp.max_row + 1):
            if r in ws_temp.row_dimensions and ws_temp.row_dimensions[r].height is not None:
                ws_out.row_dimensions[current_out_row].height = ws_temp.row_dimensions[r].height
            for c in range(1, ws_temp.max_column + 1):
                copy_cell(ws_temp.cell(row=r, column=c), ws_out.cell(row=current_out_row, column=c))
                
            for merged_range in ws_temp.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                if min_row == r:
                    try:
                        ws_out.merge_cells(
                            start_row=current_out_row, 
                            start_column=min_col, 
                            end_row=current_out_row + (max_row - min_row), 
                            end_column=max_col
                        )
                    except Exception:
                        pass
            current_out_row += 1
            
    current_out_row += 2

sel_tergabung = set()
for rentang in ws_out.merged_cells.ranges:
    for baris in range(rentang.min_row, rentang.max_row + 1):
        for kolom in range(rentang.min_col, rentang.max_col + 1):
            sel_tergabung.add((baris, kolom))

lebar_spesifik = {
    'I': 39,
    'J': 39,
    'K': 28,
    'L': 37,
    'M': 28,
    'N': 37,
    'O': 37,
    'P': 30
}

for c_idx in range(1, ws_out.max_column + 1):
    huruf_kolom = get_column_letter(c_idx)
    
    if huruf_kolom in lebar_spesifik:
        ws_out.column_dimensions[huruf_kolom].width = lebar_spesifik[huruf_kolom]
        continue
        
    panjang_maksimal = 0
    
    for r_idx in range(1, ws_out.max_row + 1):
        if (r_idx, c_idx) in sel_tergabung:
            continue
            
        cell = ws_out.cell(row=r_idx, column=c_idx)
        if cell.value is not None:
            if isinstance(cell.value, (int, float)):
                nilai_teks = f"{cell.value:,.2f}"
            else:
                nilai_teks = str(cell.value)
                
            if len(nilai_teks) > panjang_maksimal:
                panjang_maksimal = len(nilai_teks)
                
    if panjang_maksimal > 0:
        lebar_sekarang = ws_out.column_dimensions[huruf_kolom].width
        if lebar_sekarang is None:
            lebar_sekarang = 37
            
        lebar_disesuaikan = panjang_maksimal + 2.5
        if lebar_disesuaikan > lebar_sekarang:
            ws_out.column_dimensions[huruf_kolom].width = lebar_disesuaikan

for indeks_baris in range(1, ws_out.max_row + 1):
    baris_ttd = False
    for indeks_kolom in range(1, ws_out.max_column + 1):
        nilai_sel = ws_out.cell(row=indeks_baris, column=indeks_kolom).value
        if isinstance(nilai_sel, str) and "TTD SALES & COLLECTOR" in nilai_sel:
            baris_ttd = True
            break
            
    if baris_ttd:
        ws_out.row_dimensions[indeks_baris].height = 120

wb_out.save('Print_AR.xlsx')
print("--> Proses selesai, file telah disimpan sebagai Print_AR.xlsx")