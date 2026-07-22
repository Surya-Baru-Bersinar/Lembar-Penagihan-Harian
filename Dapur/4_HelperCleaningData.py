import openpyxl

def bersihkan_baris(file_input, file_output):
    print(f"--> Sedang memproses file: {file_input}")
    
    try:
        wb = openpyxl.load_workbook(file_input)
        ws = wb.active
        
        rentang_merger = list(ws.merged_cells.ranges)
        for rentang in rentang_merger:
            try:
                ws.unmerge_cells(str(rentang))
            except Exception:
                pass
                
        current_nama = ""
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                nilai_sel = ws.cell(row=row, column=col).value
                if nilai_sel is not None:
                    teks_sel = str(nilai_sel).strip()
                    if teks_sel.startswith("Nama"):
                        nama_val = ws.cell(row=row, column=4).value
                        if nama_val is not None:
                            current_nama = str(nama_val).strip()
                        break
            
            hapus_baris = False
            baris_kosong = True
            for col in range(2, ws.max_column + 1):
                val = ws.cell(row=row, column=col).value
                if val is not None and str(val).strip() != "":
                    baris_kosong = False
                    teks = str(val).strip()
                    if teks.startswith("LAPORAN HASIL") or \
                       teks.startswith("Nama") or \
                       teks.startswith("Di input oleh") or \
                       teks == "No." or \
                       teks == "TOTAL TAGIHAN" or \
                       teks.startswith("TTD SALES & COLLECTOR"):
                        hapus_baris = True
                        break
                        
            if not baris_kosong and not hapus_baris and current_nama:
                ws.cell(row=row, column=1).value = current_nama

        for row in range(ws.max_row, 0, -1):
            hapus_baris = False
            baris_kosong = True
            
            for col in range(1, ws.max_column + 1):
                nilai_sel = ws.cell(row=row, column=col).value
                if nilai_sel is not None and str(nilai_sel).strip() != "":
                    baris_kosong = False
                    teks_sel = str(nilai_sel).strip()
                    
                    if teks_sel.startswith("LAPORAN HASIL") or \
                       teks_sel.startswith("Nama") or \
                       teks_sel.startswith("Di input oleh") or \
                       teks_sel == "No." or \
                       teks_sel == "TOTAL TAGIHAN" or \
                       teks_sel.startswith("TTD SALES & COLLECTOR"):
                        hapus_baris = True
                        break
                        
            if baris_kosong or hapus_baris:
                ws.delete_rows(row, 1)
                
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 18.75
                
        wb.save(file_output)
        print(f"--> Proses selesai, file telah disimpan sebagai: {file_output}")
        
    except Exception as e:
        print(f"--> Error saat memproses file: {e}")

if __name__ == "__main__":
    file_input = 'Print_AR.xlsm'
    file_output = 'Print_AR_temp.xlsx'
    bersihkan_baris(file_input, file_output)